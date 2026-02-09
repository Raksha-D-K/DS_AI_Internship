""""

file=open("sample.txt","w")
file.write("Hello, this is a file handling example.")
file.close()
file=open("sample.txt","r")
content=file.read()
print(content)
file.close()

with open("sample.txt","r") as file:
    content=file.read()
    print(content)
    
 #error handling with r\try except block   
try:
    with open("missing.txt", "r") as file:
        print(file.read())
except FileNotFoundError:
    print("File not found. Please check the filename.")
#CSV  
import csv;
with open("src/Day7/data.csv", "r") as file:
    reader = csv.reader(file)
    for row in reader:
        print(row)

#Excel
import openpyxl
workbook = openpyxl.Workbook()
sheet = workbook.active
workbook.save("sample.xlsx")
"""

from openpyxl import load_workbook

wb = load_workbook("student.xlsx")
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row[0], row[1], row[2])
